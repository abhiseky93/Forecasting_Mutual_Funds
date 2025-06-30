import streamlit as st
import requests
import pandas as pd
import numpy as np
from sklearn.ensemble import GradientBoostingRegressor
import matplotlib.pyplot as plt
from textblob import TextBlob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as ExcelImage
import os

# Create output folder for Windows local save
os.makedirs("output", exist_ok=True)

st.set_page_config(page_title="Mutual Fund AI Bot", layout="wide")
st.title("\U0001F4C8 Mutual Fund AI Bot - NAV Predictor & Sentiment Tracker")
st.caption("Developed by Abhishek Ravikant Yadav")

NEWS_API_KEY = "e4e780e3e30043aea533701e1820962e"  # Replace with your NewsAPI key

@st.cache_data
def get_all_schemes():
    res = requests.get("https://api.mfapi.in/mf").json()
    return {s['schemeName']: s['schemeCode'] for s in res if s.get("schemeCode")}

def get_nav_data(scheme_code, days=60):
    res = requests.get(f"https://api.mfapi.in/mf/{scheme_code}").json()
    df = pd.DataFrame(res['data'][:days])
    df['nav'] = df['nav'].astype(float)
    df['date'] = pd.to_datetime(df['date'], format='%d-%m-%Y')
    df.sort_values('date', inplace=True)
    return df[['date', 'nav']]

def add_technical_features(df):
    df['pct_change'] = df['nav'].pct_change()
    df['ma_5'] = df['nav'].rolling(5).mean()
    df['volatility_5'] = df['pct_change'].rolling(5).std()
    df.dropna(inplace=True)
    return df

def fetch_sentiment_news():
    res = requests.get("https://newsapi.org/v2/everything", params={
        "q": "mutual fund india",
        "apiKey": NEWS_API_KEY,
        "language": "en",
        "sortBy": "publishedAt",
        "pageSize": 5
    }).json()
    return [article['title'] for article in res.get('articles', [])]

def analyze_sentiment(news_titles):
    return sum([TextBlob(title).sentiment.polarity for title in news_titles]) / len(news_titles)

def predict_with_features(df, sentiment_score):
    df['sentiment'] = sentiment_score
    features = ['pct_change', 'ma_5', 'volatility_5', 'sentiment']
    y = df['nav'].shift(-1)
    df.dropna(inplace=True)
    model = GradientBoostingRegressor()
    model.fit(df[features][:-1], y[:-1])
    prediction = model.predict(df[features].iloc[[-1]])[0]
    return prediction, df['nav'].iloc[-1]

def save_chart(df, scheme_name, file_path):
    fig, ax = plt.subplots()
    arrows = df['nav'].diff().apply(lambda x: '↑' if x > 0 else '↓')
    ax.plot(df['date'], df['nav'], label="NAV", color='blue')
    ax.plot(df['date'], df['ma_5'], label="5-day MA", color='green')
    for i in range(1, len(df)):
        ax.text(df['date'].iloc[i], df['nav'].iloc[i], arrows.iloc[i], fontsize=8, ha='center')
    ax.set_title(scheme_name)
    ax.set_xlabel("Date")
    ax.set_ylabel("NAV")
    ax.legend()
    ax.grid(True)
    plt.tight_layout()
    fig.savefig(file_path)
    plt.close(fig)

def export_to_excel(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    ws.append(["Scheme Name", "Current NAV", "Predicted NAV", "% Change", "Recommendation", "Reason"])

    chart_paths = []
    for idx, res in enumerate(results):
        row = [res['name'], res['current'], res['predicted'], f"{res['pct']:.2f}%", res['recommend'], res['reason']]
        ws.append(row)
        for col in range(1, 7):
            cell = ws.cell(row=idx+2, column=col)
            if res['recommend'] == "BUY":
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif res['recommend'] == "SELL":
                cell.fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
            elif res['recommend'] == "HOLD":
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        chart_path = os.path.join("output", f"chart_{idx}.png")
        save_chart(res['nav_df'], res['name'], chart_path)
        chart_paths.append(chart_path)

    ws.append([])
    ws.append(["Developed by Abhishek Ravikant Yadav"])

    for i, path in enumerate(chart_paths):
        img = ExcelImage(path)
        ws.add_image(img, f"A{len(results) + 4 + (i * 20)}")

    file_path = os.path.join("output", "MutualFundReport.xlsx")
    wb.save(file_path)
    return file_path

schemes = get_all_schemes()
selected_names = st.multiselect("\U0001F50D Select Mutual Funds", sorted(schemes.keys()), max_selections=5)

if selected_names:
    news = fetch_sentiment_news()
    sentiment = analyze_sentiment(news)

    st.markdown("""
    <marquee behavior="scroll" direction="left" scrollamount="5" style="color:blue; font-weight:bold; background:#f0f2f6; padding:8px; border-radius:5px;">
    {}
    </marquee>
    """.format(" | ".join(news)), unsafe_allow_html=True)

    st.markdown(f"**Average Sentiment Score:** `{sentiment:.2f}`")

    results = []

    for name in selected_names:
        code = schemes[name]
        df_raw = get_nav_data(code)
        df_feat = add_technical_features(df_raw.copy())
        pred, curr = predict_with_features(df_feat.copy(), sentiment)
        pct = ((pred - curr) / curr) * 100
        if pct > 2 and sentiment > 0:
            rec = "BUY"
            reason = "Predicted NAV growth and positive market sentiment"
        elif pct < -2 and sentiment < 0:
            rec = "SELL"
            reason = "Predicted decline and negative sentiment"
        else:
            rec = "HOLD"
            reason = "Stable NAV prediction with neutral sentiment"
        results.append({
            'name': name,
            'predicted': round(pred, 2),
            'current': round(curr, 2),
            'pct': pct,
            'recommend': rec,
            'reason': reason,
            'nav_df': df_feat
        })

        with st.expander(f"\U0001F4C5 30-Day NAV Movement for {name}"):
            df_trend = df_raw.tail(30).copy()
            df_trend['prev_nav'] = df_trend['nav'].shift(1)
            df_trend['change'] = df_trend['nav'] - df_trend['prev_nav']
            df_trend['direction'] = df_trend['change'].apply(lambda x: "\U0001F53C" if x > 0 else ("\U0001F53D" if x < 0 else "➖"))
            df_trend.dropna(inplace=True)

            styled_df = df_trend[['date', 'nav', 'direction']].rename(columns={
                'date': 'Date', 'nav': 'NAV', 'direction': 'Up/Down'
            }).style.apply(
                lambda row: ['background-color: #d4edda' if row['Up/Down'] == '🔼'
                             else 'background-color: #f8d7da' if row['Up/Down'] == '🔽'
                             else '' for _ in row], axis=1)

            st.dataframe(styled_df, use_container_width=True)

            up_days = df_trend['direction'].value_counts().get("🔼", 0)
            down_days = df_trend['direction'].value_counts().get("🔽", 0)
            st.markdown(f"**🔼 Up Days:** {up_days} &nbsp;&nbsp; **🔽 Down Days:** {down_days}")

    st.subheader("\U0001F4CA Predictions")
    df_summary = pd.DataFrame([{
        "Scheme": r['name'],
        "Current NAV": r['current'],
        "Predicted NAV": r['predicted'],
        "% Change": f"{r['pct']:.2f}%",
        "Recommendation": r['recommend'],
        "Reason": r['reason']
    } for r in results])
    st.dataframe(df_summary)

    if st.button("\U0001F4E5 Export to Excel"):
        excel_path = export_to_excel(results)
        with open(excel_path, "rb") as f:
            st.download_button("Download Excel File", f, file_name="MutualFundReport.xlsx")
